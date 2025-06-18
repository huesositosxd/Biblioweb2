from django.core.management.base import BaseCommand
from django.utils.timezone import now
from django.core.mail import EmailMessage
from django.conf import settings
from Aplicaciones.Visitas.models import RegistroVisita, Visitausuario, RegistroUsuario

import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import os


class Command(BaseCommand):
    help = 'Genera y envía reporte diario de visitas'

    def handle(self, *args, **kwargs):
        hoy = now().date()

        # --- ELIMINAR USUARIOS NO VALIDADOS ---
        usuarios_no_validados = RegistroUsuario.objects.filter(verificado=0)
        eliminados = usuarios_no_validados.count()
        usuarios_no_validados.delete()
        self.stdout.write(self.style.WARNING(f"{eliminados} usuarios no validados eliminados."))

        # --- OBTENER VISITAS ANÓNIMAS ---
        visitas_anonimas = RegistroVisita.objects.filter(fecha_hora__date=hoy)
        df_anonimas = pd.DataFrame(list(visitas_anonimas.values()))
        if not df_anonimas.empty:
            df_anonimas["tipo_visita"] = "anónima"
            df_anonimas.rename(columns={"fecha_hora": "fecha_visita"}, inplace=True)

        # --- OBTENER VISITAS AUTENTICADAS ---
        visitas_usuarios = Visitausuario.objects.filter(fecha_visita__date=hoy)
        df_usuarios = pd.DataFrame(list(visitas_usuarios.values()))
        if not df_usuarios.empty:
            df_usuarios["tipo_visita"] = "usuario"
            if "fecha_visita" not in df_usuarios.columns:
                for col in df_usuarios.columns:
                    if col.lower() == "fecha_visita":
                        df_usuarios.rename(columns={col: "fecha_visita"}, inplace=True)

        # --- CAMPOS COMUNES ---
        campos_comunes = [
            "nombre", "fecha_nacimiento", "edad", "sexo",
            "colonia", "discapacidad", "actividad", "estado_nacimiento",
            "fecha_visita", "tipo_visita", "parque"
        ]
        df_anonimas = df_anonimas[[c for c in campos_comunes if c in df_anonimas.columns]]
        df_usuarios = df_usuarios[[c for c in campos_comunes if c in df_usuarios.columns]]

        # --- UNIFICAR VISITAS ---
        df_final = pd.concat([df_anonimas, df_usuarios], ignore_index=True)

        if "fecha_visita" in df_final.columns and pd.api.types.is_datetime64tz_dtype(df_final["fecha_visita"]):
            df_final["fecha_visita"] = df_final["fecha_visita"].dt.tz_localize(None)

        # --- PARQUES PARA HOJAS DE EXCEL ---
        parques = [
            "Gimnasio Municipal", "BiblioParque Norte", "BiblioParque sur",
            "Multideportivo el Sarape", "Unidad Deportiva Carlos R. Gonzáles",
            "Unidad Deportiva Jesús Carranza", "Parque los Nogales",
            "Unidad Benito Juárez", "Alberca Municipal Saltillo 2000"
        ]

        # --- GENERAR EXCEL CON HOJAS POR PARQUE ---
        nombre_archivo = f"reporte_visitas_{hoy}.xlsx"
        ruta = os.path.join(settings.BASE_DIR, nombre_archivo)

        with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
            for parque in parques:
                df_parque = df_final[df_final["parque"] == parque]
                if not df_parque.empty:
                    hoja_nombre = parque[:31]
                    df_parque.to_excel(writer, sheet_name=hoja_nombre, index=False)
            df_final.to_excel(writer, sheet_name="Resumen General", index=False)

        # --- DAR FORMATO A HOJAS DE EXCEL ---
        wb = openpyxl.load_workbook(ruta)
        for sheet in wb.worksheets:
            max_col = sheet.max_column
            max_row = sheet.max_row

            if max_row < 2 or max_col < 1:
                continue

            # Ajustar ancho de columnas
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[column].width = max_length + 2

            # Crear tabla visual
            tab_ref = f"A1:{sheet.cell(row=1, column=max_col).column_letter}{max_row}"
            table = Table(displayName=f"Tabla_{sheet.title.replace(' ', '_')}", ref=tab_ref)

            style = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            table.tableStyleInfo = style
            sheet.add_table(table)

        wb.save(ruta)

        # --- ENVIAR CORREO CON REPORTE ---
        email = EmailMessage(
            subject="Reporte Diario de Visitas",
            body="Se adjunta el reporte de visitas del día de hoy.",
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=["josueegremi@gmail.com"],
        )
        email.attach_file(ruta)
        email.send()

        # --- ELIMINAR ARCHIVO LOCAL ---
        try:
            os.remove(ruta)
            self.stdout.write(self.style.SUCCESS(f"Archivo eliminado: {nombre_archivo}"))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"No se pudo eliminar el archivo: {e}"))

        self.stdout.write(self.style.SUCCESS(f"Reporte generado y enviado correctamente: {nombre_archivo}"))
